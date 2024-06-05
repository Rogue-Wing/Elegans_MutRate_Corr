# Takes the Konrad data (as a .JSON file) and the Excel Spreadsheet of Saxena data, and makes two .json files of mutations (one for each)

import openpyxl as pyx
import json
import os

PATH = f'{os.path.dirname(__file__)}/Resources'.replace('\\', '/') # The path to the 'Resources' folder of the project
MUTATION_LIST_1 = f'{PATH}/C. Elegans Mutations (Saxena, A. S. et al).xlsx' # The path (from Resources) to the list of mutations in two subsequent mutation lines, presented in Saxena, A. S. et al (2008) 
MUTATION_LIST_2 = f'{PATH}/Konrad A. et al CE Data.json'

def testForDuplicateLoci(mutations1, mutations2): # Iterates through each element in one list of (formatted) mutations, comparing their loci to all the loci in a second list. If a loci already appears in the second list, the mutation is not added again.
    for mutation1 in mutations1:
        duplicate = False
        for mutation2 in mutations2:
            if mutation1['locus'] == mutation2['locus']:
                duplicate = True
                print(mutation1['locus'])
        if not duplicate:
            mutations2.append(mutation1)
    
    return mutations2

def convertNumeralToNumber(numeral):
    numeralToNum = {'I':1, 'II':2, 'III':3, 'IV':4, 'V':5, 'VI':6, 'VII':7, 'VIII':8, 'IX':9, 'X':10, 'XI':11, 'XII':12, 'MtDNA':13}
    return numeralToNum[numeral]

MutationSheets = pyx.load_workbook(MUTATION_LIST_1) # Loads the Excel spreadsheet into Python

#! NOTE: Duplications are being allowed - mutations.append have been changed to allMutations.append

def getSaxenaMutations(mutationSheets):
    allMutations = []
    SoI = mutationSheets['O2_SNP'] # Gets the reference to the specific sheet in the spreadsheet for the O2MA mutation list
    for row in range(2, SoI.max_row + 1): #! For each entry in the O2MA mutation list (48 lines sequenced), compile the locus (<CHROMOSOME NUMBER>:<POSITION>), gene name, SNP and type of mutation together into a dict 
        locus = f"{SoI.cell(row, 2).value.split('_')[1]}:{SoI.cell(row, 3).value}"
        gene = SoI.cell(row, 7).value
        substitution = f"{SoI.cell(row, 4).value} -> {SoI.cell(row, 5).value}"
        mutationType = SoI.cell(row, 6).value 
        allMutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(SoI.cell(row, 2).value.split('_')[1]), int(SoI.cell(row, 3).value))}) # Add the information to the values list

    SoI = mutationSheets['O1MA_Ancestor_SNP'] # Gets the reference to the specific sheet in the spreadsheet for the O1MA ancestors (those that the O2MA lines came from) mutation list
    mutations = []
    for row in range(2, SoI.max_row + 1): #! Repeat for O1MA (9 lines sequenced)
        locus = f"{SoI.cell(row, 2).value.split('_')[1]}:{SoI.cell(row, 3).value}"
        gene = SoI.cell(row, 7).value
        substitution = f"{SoI.cell(row, 4).value} -> {SoI.cell(row, 5).value}"
        mutationType = SoI.cell(row, 6).value 
        mutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(SoI.cell(row, 2).value.split('_')[1]), int(SoI.cell(row, 3).value))})

    allMutations = testForDuplicateLoci(mutations, allMutations)

    SoI = mutationSheets['O1MA_Other_SNP'] # Gets the reference to the specific sheet in the spreadsheet for the O1MA (non-ancestors) mutation list
    mutations = []
    for row in range(2, SoI.max_row + 1): #! Repeat for the other O1MA lines that were sequenced (23 in total). Note that the sheet's 'Chromosome' column is formatted slightly differently
        locus = f"{SoI.cell(row, 2).value}:{SoI.cell(row, 3).value}"
        gene = SoI.cell(row, 7).value
        substitution = f"{SoI.cell(row, 4).value} -> {SoI.cell(row, 5).value}"
        mutationType = SoI.cell(row, 6).value 
        mutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(SoI.cell(row, 2).value), int(SoI.cell(row, 3).value))})

    allMutations = testForDuplicateLoci(mutations, allMutations)

    SoI = mutationSheets['O2_Mononuc_SNP'] # Gets the reference to the specific sheet in the spreadsheet for the O2MA mononucleotide STR SNPs mutation list
    mutations = []
    for row in range(2, SoI.max_row + 1): #! Repeat for the other O2MA mononucleotide STR-specific SNPs (45 lines). Note the altered 'Chromosome', as well as the lack of a gene name or mutation type
        locus = f"{SoI.cell(row, 2).value[3:]}:{SoI.cell(row, 3).value}"
        gene = None
        substitution = f"{SoI.cell(row, 8).value} -> {SoI.cell(row, 9).value}"
        mutationType = None 
        mutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(SoI.cell(row, 2).value[3:]), int(SoI.cell(row, 3).value))})

    allMutations = testForDuplicateLoci(mutations, allMutations)

    SoI = mutationSheets['O2_Dinuc_SNP'] # Gets the reference to the specific sheet in the spreadsheet for the O2MA dinucleotide STR SNPs mutation list
    mutations = []
    for row in range(2, SoI.max_row + 1): #! Repeat for the other O2MA dinucleotide STR-specific SNPs (8 lines). Note the altered 'Chromosome', as well as the lack of a gene name or mutation type
        locus = f"{SoI.cell(row, 2).value[3:]}:{SoI.cell(row, 3).value}"
        gene = None
        substitution = f"{SoI.cell(row, 8).value} -> {SoI.cell(row, 9).value}"
        mutationType = None 
        mutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(SoI.cell(row, 2).value[3:]), int(SoI.cell(row, 3).value))})

    allMutations = testForDuplicateLoci(mutations, allMutations)

    allMutations.sort(key = lambda x: x['orderingData']) # Sort the list of mutations in ascending order by locus

    print(len(allMutations))

    return allMutations

def getKonradMutations():
    mutations = []
    with open(MUTATION_LIST_2, 'r') as FoI: # Opens the .JSON file of formatted Konrad mutations. Note that this occurs before the Mono/Dinuc sheets because these hold more information, so should be prioritised in the dupe check
        for line in FoI:
            data = json.loads(line.strip())
            locus = f"{data['Chromosome']}:{data['Position']}"
            gene = None
            substitution = data['Mutation']
            mutationType = f"{data['CT']}/{data['Effect']}" # Gets the mutationType as <GENE_TYPE>/<EFFECT>, where GENE_TYPE can be an intron, exon or IG, and the effect can be synonymous, a subsitution or - (for introns/IGs)
            mutations.append({'locus':locus, 'gene':gene, 'substitution':substitution, 'mutationType':mutationType, 'orderingData':(convertNumeralToNumber(data['Chromosome']), int(data['Position']))})
    
    mutations = testForDuplicateLoci(mutations, [])

    mutations.sort(key = lambda x: x['orderingData']) # Sort the list of mutations in ascending order by locus

    print(len(mutations))

    return mutations

saxenaMutants = getSaxenaMutations(MutationSheets)
konradMutants = getKonradMutations()

with open(f'{PATH}/Konrad Unfiltered Data.json', 'w') as f:
    for mutation in konradMutants:
        f.write(json.dumps(mutation) + '\n')

#? This is the code that was used to format the Konrad data from its initial .PDF format, manually copied into a text file, into a more manipulatable (and computer-readable) .JSON format

#! This is to take the raw input of the .pdf (with the relavant data copied over to a text file and the page numbers/n counts removed), and convert it into a .json of mutations 
'''
mutations = []
data = []
with open(f'{PATH}/Konrad A. et al CE Raw.json', 'r') as f:
    for line in f:
        line = line.strip()
        data += line.split(' ')

clear = False
while not clear:
    clear = True
    for i, dataPoint in enumerate(data):
        if dataPoint == '-â€º':
            clear = False
            data[i] = data[i - 1] + ' -> ' + data[i + 1]
            data.pop(i + 1)
            data.pop(i - 1)
            break

while len(data) >= 9:
    mutation = []
    for i in range(9):
        mutation.append(data.pop(0))
    mutations.append(mutation)

with open(f'{PATH}/Konrad A. et al CE Data Formatted.json', 'w') as f:
    for mutation in mutations:
        f.write(json.dumps({'Line':mutation[0], 'Chromosome':mutation[1], 'Position':mutation[2], 'Mutation':mutation[3], 'Frequency':mutation[4], 'CT':mutation[5], 'Effect':mutation[6], 'RZ':mutation[7], 'GE':mutation[8]}) + '\n')
'''

#! This removes the commas from the Konrad Position data
'''
mutations = []
with open(XYZ, 'r') as FoI:
    for line in FoI:
        data = json.loads(line.strip())
        data['Position'] = ''.join(data['Position'].split(','))
        mutations.append(data)

with open(XYZ, 'w') as FoI:
    for mutation in mutations:
        FoI.write(json.dumps({'Line':mutation['Line'], 'Chromosome':mutation['Chromosome'], 'Position':mutation['Position'], 'Mutation':mutation['Mutation'], 'Frequency':mutation['Frequency'], 'CT':mutation['CT'], 'Effect':mutation['Effect'], 'RZ':mutation['RZ'], 'GE':mutation['GE']}) + '\n')
'''