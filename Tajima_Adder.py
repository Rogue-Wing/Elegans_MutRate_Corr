# Takes a complied .JSON file of mutants (with dN/dS ratios) and adds their respective Tajima D value from 'TajD_allgenes.xlsx'

import json
import openpyxl as pyx
import requests
import os

PATH = f'{os.path.dirname(__file__)}/Resources'.replace('\\', '/') # The path to the 'Resources' folder of the project

tajDSheet = pyx.load_workbook(f'{PATH}/TajD_allgenes.xlsx')['in'] # The Excel spreadsheet that contains the TajD and pi data

def progressBar(currentPercent, tick): # Creates a string for the progress bar
    symbol = {0:'|', 1:'/', 2:'-', 3:'\\'}[tick]

    tick += 1
    if tick > 3:
        tick = 0

    return f"{'-' * (currentPercent)}{' ' * (100 - currentPercent)} [{currentPercent}%] {symbol}\r", tick

stats = {}
numRows = tajDSheet.max_row
for row in range(2, numRows + 1): # Read in all the data from the spreadsheet
    WBID = tajDSheet.cell(row, 1).value
    tajDValue = tajDSheet.cell(row, 4).value
    if tajDValue == 'NA':
        tajDValue = None
    piValue = tajDSheet.cell(row, 5).value
    stats[WBID] = {'tajD':tajDValue, 'pi':piValue}

with open(f'{PATH}/Combined NS-MA Raw Data v3.json', 'r') as FoI: # Open the combined file of MA and AG data
    mutants = json.loads(FoI.read())

mutsPerPercent = len(mutants) // 100
bar, tick = progressBar(0, 0)
print(bar, end = '')

errs = 0
for mutNum, mutant in enumerate(mutants): # For every gene in the database
    try:
        mutants[mutant]['tajD'] = stats[mutant]['tajD'] # Try to add the Tajima's D and pi values to the data
        mutants[mutant]['piNorm'] = stats[mutant]['pi']
    except KeyError: # If this doesn't work, try and get an updated WBID for the gene
        merged = requests.get(f"http://api.wormbase.org//rest/field/gene/{mutant}/merged_into").json()['merged_into']['data']
        if merged != None:
            try:
                mutants[mutant]['tajD'] = stats[merged['id']]['tajD'] # And then try to use this updated WBID
                mutants[mutant]['piNorm'] = stats[merged['id']]['pi']
            except KeyError: # If this doesn't work, there's no data for that gene, so = None
                errs += 1
                mutants[mutant]['tajD'] = None
                mutants[mutant]['piNorm'] = None
                #- geneType = requests.get(f"http://api.wormbase.org//rest/field/gene/{merged['id']}/classification").json()['classification']['data']['type']
            #- print(merged)
        else: # If this doesn't work, there's no data for that gene, so = None
            errs += 1
            mutants[mutant]['tajD'] = None
            mutants[mutant]['piNorm'] = None
            #- geneType = requests.get(f"http://api.wormbase.org//rest/field/gene/{mutant}/classification").json()['classification']['data']['type']
            #- print(f'{mutant} has no associated tajD data!')
    
    bar, tick = progressBar(mutNum // mutsPerPercent, tick)
    print(bar, end = '')

print(f'Done! {" " * 120}')

print(errs)

with open(f'{PATH}/Combined NS-MA-TajD Raw Data v3.json', 'w') as FoI: # Save the data
    FoI.write(json.dumps(mutants))